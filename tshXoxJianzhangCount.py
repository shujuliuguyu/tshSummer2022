import openpyxl
import requests
xlsx = openpyxl.load_workbook(
    r'C:\Users\acer\Desktop\tshXoxZhibijianInfo.xlsx')
sheet = xlsx.worksheets[0]
number = sheet.max_row
url1 = "https://api.live.bilibili.com/xlive/app-room/v2/guardTab/topList?roomid="
url2 = "&page="
url3 = "&ruid="
url4 = "&page_size=29"
rList = []

for i in range(2, number+1):
    r = requests.get(url1+str(sheet.cell(i, 3).value)+url2 +
                     str(1)+url3+str(sheet.cell(i, 4).value)+url4)
    rList.append(r)

dList = []
for i in rList:
    dList.append(i.json())

for i in dList:
    print(str(i['data']['info']['num'])+" ", end="")  # 输出大航海总数
    zongdu = 0
    tidu = 0
    for j in i['data']['list']:
        if(j['guard_level'] == 1):
            zongdu += 1
        if(j['guard_level'] == 2):
            tidu += 1
    for j in i['data']['top3']:
        if(j['guard_level'] == 1):
            zongdu += 1
        if(j['guard_level'] == 2):
            tidu += 1
    print(str(zongdu)+" "+str(tidu))  # 输出总督和提督数量
